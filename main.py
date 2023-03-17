import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
products_list =inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier={}
products_under_10_inv = {}


for product_row in range(2,products_list.max_row + 1):
    supplier_name=products_list.cell(product_row, 4).value
    inventory =products_list.cell(product_row, 2).value
    price =products_list.cell(product_row, 3).value
    products_num =products_list.cell(product_row, 1).value
    #calculo numero de productos por supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] =products_per_supplier[supplier_name]+ 1
    else:
        products_per_supplier[supplier_name] = 1
    #calcular el valor total de inventorio por supplier

    if supplier_name in total_value_per_supplier:
        current_total_value =total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name]= inventory * price

    #calcular productos con inventario menos de 10
    if inventory < 10:
        products_under_10_inv[int(products_num)]= int(inventory)


print(products_under_10_inv)

